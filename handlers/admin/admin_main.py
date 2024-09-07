from aiogram import types, Router, F, Bot
import pandas as pd
from datetime import timedelta

from aiogram.types import FSInputFile
from aiogram.filters import CommandStart, Command
from aiogram.types import CallbackQuery, InlineKeyboardButton, Message
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from config_data.config import Config, load_config
from filters.chat_types import ChatTypeFilter
from filters.is_admin import IsAdminMsg
from keyboards.inline.inline import get_callback_btns, get_inlineMix_btns
from lexicon.lexicon import LEXICON_btn_main_admin_menu, LEXICON_ADMIN, LEXICON_btn_BACK_main_admin_menu
from database.orm_query import orm_get_users
from database.models import (User, CourseRequest, CodeMissin, BadCode, WhereEnterCode,
                             CanNotEnterAccaunt, NoQuestion)
from aiogram.filters import Command, or_f
from dotenv import find_dotenv, load_dotenv

from sqlalchemy.future import select

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment



load_dotenv(find_dotenv())
config: Config = load_config()

admin_router = Router()
admin_router.message.filter(ChatTypeFilter(['private']), IsAdminMsg())
admin_router.callback_query.filter(IsAdminMsg())


@admin_router.message(Command("admin"), F.text | F.command)
@admin_router.message(F.text.lower().in_({'admin', 'administrator', "админ", "администратор"}))
async def admin_handler_message(message: Message):
    await message.answer(text=LEXICON_ADMIN["/admin+panel"],
                         reply_markup=get_callback_btns(btns=LEXICON_btn_main_admin_menu, sizes=(2,)))


@admin_router.callback_query(F.data == "admin_menu")
async def admin_handler_caccback(callback: CallbackQuery):
    await callback.message.answer(text=LEXICON_ADMIN["/admin+panel"],
                                  reply_markup=get_callback_btns(btns=LEXICON_btn_main_admin_menu, sizes=(2,)))


@admin_router.callback_query(F.data == "instruction_admin")
async def get_admin_instruction(callback: CallbackQuery):
    await callback.message.answer(text=LEXICON_ADMIN["/instruction_description"],
                                  reply_markup=get_callback_btns(btns=LEXICON_btn_BACK_main_admin_menu))


@admin_router.callback_query(F.data == "users_all_list")
async def get_users_list(callback: CallbackQuery, session: AsyncSession):
    # Получение всех пользователей из базы данных
    result = await session.execute(select(User))
    users = result.scalars().all()

    # Создание списка словарей, где ключи соответствуют столбцам
    data = [
        {
            "ID": user.id,
            "User ID": user.user_id,
            "ИМЯ": user.first_name,
            "ФАМИЛИЯ": user.last_name,
            "ЛОГИН": user.username,
            "НОМЕР ТЕЛ./Phone": user.phone
        }
        for user in users
    ]

    # Создание DataFrame из списка словарей
    df = pd.DataFrame(data)

    # Указание имени файла
    file_name = "users_data.xlsx"


    # Максимальная ширина столбца (например, 30 символов)
    MAX_COLUMN_WIDTH = 30

    # Сохранение DataFrame в Excel-файл с изменением ширины столбцов и переносом текста
    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        worksheet = writer.sheets['Sheet1']

        # Настройка ширины столбцов и перенос текста
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column_cells)
            adjusted_width = min(max_length + 2, MAX_COLUMN_WIDTH)  # Ограничение ширины
            column_letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Перенос текста для всех ячеек
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

    # Создание InputFile с использованием пути к файлу
    input_file = FSInputFile(file_name)

    # Отправка файла
    await callback.message.answer_document(input_file)


@admin_router.callback_query(F.data == "report_users")
async def get_admin_report(callback: CallbackQuery, session: AsyncSession):
    # Получение данных по проблеме подбор курса из базы данных
    result_help_with_coure = await session.execute(select(CourseRequest))
    course_requests = result_help_with_coure.scalars().all()

    # Получение данных по проблеме не пришел код из базы данных
    result_code_missing = await session.execute(select(CodeMissin))
    code_missing_request = result_code_missing.scalars().all()

    # Получение данных по проблеме не понимаю куда вводить код
    result_where_entercode = await session.execute(select(WhereEnterCode))
    where_entercod_request = result_where_entercode.scalars().all()

    # Получение данных по проблеме пришел код но он не работает
    result_badcode = await session.execute(select(BadCode))
    badcode_request = result_badcode.scalars().all()

    # Получение данных по проблеме не могу войти в аккаунт
    result_cannot_enter = await session.execute(select(CanNotEnterAccaunt))
    cannot_enter_request = result_cannot_enter.scalars().all()

    # Получение данных по проблеме моего вопроса нет в списке
    result_noquestion = await session.execute(select(NoQuestion))
    noquestion_request = result_noquestion.scalars().all()

    # Функция для добавления +3 часов к UTC времени
    def adjust_time(created_time):
        return created_time + timedelta(hours=3)


    # Создание списка словарей, где ключи соответствуют столбцам
    data_course = [
        {
            "Дата создания": adjust_time(req.created),  # Добавляем +3 часа
            "Проблема": "Помощь с подбором курса",
            "User ID": req.user_id,
            "ИМЯ": req.first_name,
            "ФАМИЛИЯ": req.last_name,
            "ЛОГИН": req.username,
            "Ответ1": req.question1,
            "Ответ2": req.question2,
            "Ответ3": req.question3,
        }
        for req in course_requests
    ]

    data_code_missing = [
        {
            "Дата создания": adjust_time(req.created),  # Добавляем +3 часа
            "Проблема": "Не пришел код",
            "User ID": req.user_id,
            "ИМЯ": req.first_name,
            "ФАМИЛИЯ": req.last_name,
            "ЛОГИН": req.username,
            "Почта": req.mail_user,
        }
        for req in code_missing_request
    ]

    data_where_entercode = [
        {
            "Дата создания": adjust_time(req.created),  # Добавляем +3 часа
            "Проблема": "Куда вводить код",
            "User ID": req.user_id,
            "ИМЯ": req.first_name,
            "ФАМИЛИЯ": req.last_name,
            "ЛОГИН": req.username,
        }
        for req in where_entercod_request
    ]

    data_badcode = [
        {
            "Дата создания": adjust_time(req.created),  # Добавляем +3 часа
            "Проблема": "Не работает код",
            "User ID": req.user_id,
            "ИМЯ": req.first_name,
            "ФАМИЛИЯ": req.last_name,
            "ЛОГИН": req.username,
        }
        for req in badcode_request
    ]

    data_cannot_enter = [
        {
            "Дата создания": adjust_time(req.created),  # Добавляем +3 часа
            "Проблема": "Не могу войти в аккаунт",
            "User ID": req.user_id,
            "ИМЯ": req.first_name,
            "ФАМИЛИЯ": req.last_name,
            "ЛОГИН": req.username,
            "Почта": req.mail_user,
        }
        for req in cannot_enter_request
    ]

    data_noquestion = [
        {
            "Дата создания": adjust_time(req.created),  # Добавляем +3 часа
            "Проблема": "Нет моего вопроса",
            "User ID": req.user_id,
            "ИМЯ": req.first_name,
            "ФАМИЛИЯ": req.last_name,
            "ЛОГИН": req.username,
            "Доп.вопрос": req.question_user,
        }
        for req in noquestion_request
    ]

    combined_data = (data_course + data_code_missing + data_where_entercode +
                     data_badcode + data_cannot_enter + data_noquestion)

    # Создание DataFrame из списка словарей
    df = pd.DataFrame(combined_data)

    # Сортировка данных по полю 'Дата создания' в порядке убывания (сначала свежие)
    df = df.sort_values(by="Дата создания", ascending=False)

    # Указание имени файла
    file_name = "report.xlsx"

    # Максимальная ширина столбца (например, 30 символов)
    MAX_COLUMN_WIDTH = 30

    # Сохранение DataFrame в Excel-файл с изменением ширины столбцов и переносом текста
    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        worksheet = writer.sheets['Sheet1']

        # Настройка ширины столбцов и перенос текста
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column_cells)
            adjusted_width = min(max_length + 2, MAX_COLUMN_WIDTH)  # Ограничение ширины
            column_letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Перенос текста для всех ячеек
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

    # Создание InputFile с использованием пути к файлу
    input_file = FSInputFile(file_name)

    # Отправка файла
    await callback.message.answer_document(input_file)
